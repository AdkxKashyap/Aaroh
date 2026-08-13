"""Utilities for extracting readable text from uploaded document files."""

from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path

import pymupdf
from openpyxl import load_workbook


class ExtractionAdapter:
    """Simple extraction adapter for the current POC document pipeline.

    This intentionally keeps the boundary narrow: the document service reads the
    stored bytes, and this adapter turns those bytes into text for the parser.
    The current implementation supports plain text and simple PDF extraction.
    """

    _TEXT_EXTENSIONS = {".txt", ".md", ".csv", ".json"}
    _SUPPORTED_EXTENSIONS = _TEXT_EXTENSIONS | {".pdf", ".xlsx"}

    @classmethod
    def prepare_parser_input(
        cls, file_bytes: bytes, filename: str | None = None
    ) -> str:
        """Return a normalized text payload intended for the parser boundary."""
        extracted = cls.extract_text(file_bytes, filename)
        normalized = re.sub(r"\s+", " ", extracted).strip()
        if not normalized:
            raise ValueError("Document content is empty or unreadable.")
        return normalized

    @classmethod
    def extract_text(cls, file_bytes: bytes, filename: str | None = None) -> str:
        if not file_bytes:
            raise ValueError("Document content is empty.")

        suffix = (Path(filename or "").suffix or "").lower()

        if suffix in cls._TEXT_EXTENSIONS or suffix == "":
            return cls._extract_text_bytes(file_bytes)

        if suffix == ".pdf":
            return cls._extract_pdf_text(file_bytes)

        if suffix == ".xlsx":
            return cls._extract_xlsx_text(file_bytes)

        raise ValueError(f"Unsupported document type: {suffix or 'unknown'}")

    @staticmethod
    def _extract_text_bytes(file_bytes: bytes) -> str:
        try:
            text = file_bytes.decode("utf-8")
        except UnicodeDecodeError:
            text = file_bytes.decode("utf-8", errors="replace")

        cleaned = re.sub(r"\s+", " ", text).strip()
        if not cleaned:
            raise ValueError("Document content is empty or unreadable.")
        return cleaned

    @staticmethod
    def _extract_pdf_text(file_bytes: bytes) -> str:
        if pymupdf is None:
            raise ValueError(
                "PyMuPDF is required to extract PDF text. Install PyMuPDF in the app environment."
            )

        try:
            with pymupdf.open(stream=file_bytes, filetype="pdf") as document:
                pages: list[str] = []
                for page in document:
                    text = page.get_text("text")
                    if text and text.strip():
                        pages.append(text.strip())
                if pages:
                    return "\n".join(pages)
        except Exception as exc:
            raise ValueError(
                "PDF content could not be extracted with PyMuPDF."
            ) from exc

        raise ValueError("No readable text was found in the PDF.")

    @staticmethod
    def _extract_xlsx_text(file_bytes: bytes) -> str:
        workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                values = ["" if cell is None else str(cell).strip() for cell in row]
                if any(values):
                    lines.append(",".join(values))

        text = "\n".join(lines).strip()
        if not text:
            raise ValueError("Spreadsheet content is empty or unreadable.")
        return text
